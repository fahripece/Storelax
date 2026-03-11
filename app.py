import sqlite3, json, os, io, csv, time, threading
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, jsonify, g,
                   session, redirect, url_for, send_file)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, "templates"),
            static_folder=os.path.join(BASE_DIR, "static"))
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production-please")
DB_PATH = os.path.join(BASE_DIR, "inventory.db")

# ── Security headers ────────────────────────────────────────────────────────────
@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response

# ── Global error handlers ───────────────────────────────────────────────────────
@app.errorhandler(404)
def err_404(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok":False,"msg":"Not found"}), 404
    return redirect(url_for("inventory"))

@app.errorhandler(500)
def err_500(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok":False,"msg":"Internal server error"}), 500
    return render_template("login.html", error="An unexpected error occurred."), 500

# ── Login rate limiter (in-memory, per IP) ──────────────────────────────────────
_login_attempts = {}   # {ip: [timestamp, ...]}
_login_lock     = threading.Lock()
LOGIN_WINDOW    = 300  # 5 minutes
LOGIN_MAX       = 10   # max attempts per window

def check_login_rate(ip):
    """Returns (allowed, seconds_until_reset). Cleans old entries."""
    now = time.time()
    with _login_lock:
        attempts = [t for t in _login_attempts.get(ip, []) if now - t < LOGIN_WINDOW]
        _login_attempts[ip] = attempts
        if len(attempts) >= LOGIN_MAX:
            reset_in = int(LOGIN_WINDOW - (now - attempts[0]))
            return False, reset_in
        _login_attempts[ip] = attempts + [now]
        return True, 0

def clear_login_rate(ip):
    with _login_lock:
        _login_attempts.pop(ip, None)

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
        g.db.execute("PRAGMA synchronous=NORMAL")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db: db.close()

def query(sql, args=(), one=False):
    cur = get_db().execute(sql, args)
    rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    db = get_db()
    cur = db.execute(sql, args)
    db.commit()
    return cur.lastrowid

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT UNIQUE NOT NULL,
            password    TEXT NOT NULL,
            role        TEXT NOT NULL DEFAULT 'worker',
            permissions TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS categories (
            id    INTEGER PRIMARY KEY AUTOINCREMENT,
            name  TEXT UNIQUE NOT NULL,
            color TEXT NOT NULL DEFAULT '#1d4ed8'
        );
        CREATE TABLE IF NOT EXISTS product_types (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            name                  TEXT UNIQUE NOT NULL,
            serial_required       INTEGER DEFAULT 0,
            sku_required          INTEGER DEFAULT 0,
            qty_tracked           INTEGER DEFAULT 0,
            require_scan_checkout INTEGER DEFAULT 0,
            active                INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS item_modifications (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id         INTEGER NOT NULL REFERENCES items(id),
            ts              TEXT NOT NULL,
            modified_by     TEXT NOT NULL,
            field_changed   TEXT NOT NULL,
            old_value       TEXT,
            new_value       TEXT,
            notes           TEXT,
            spawned_item_id INTEGER REFERENCES items(id)
        );
        CREATE TABLE IF NOT EXISTS items (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            name                TEXT NOT NULL,
            manufacturer        TEXT,
            model               TEXT,
            serial              TEXT,
            sku                 TEXT,
            category_id         INTEGER REFERENCES categories(id),
            product_type_id     INTEGER REFERENCES product_types(id),
            condition           TEXT DEFAULT 'New',
            owner_company       TEXT,
            purchase_date       TEXT,
            shelf               TEXT,
            qty                 INTEGER,
            qty_out             INTEGER DEFAULT 0,
            low_stock_threshold INTEGER DEFAULT 0,
            checked_out         INTEGER DEFAULT 0,
            checkout_date       TEXT,
            checkout_by         TEXT,
            job_ref             TEXT,
            notes               TEXT,
            cpu                 TEXT,
            ram                 TEXT,
            storage             TEXT,
            cost_price          REAL,
            sale_price          REAL,
            tax_paid            INTEGER DEFAULT -1,
            tax_rate            REAL DEFAULT 0,
            sold                INTEGER DEFAULT 0,
            sold_date           TEXT,
            sold_price          REAL,
            sold_to             TEXT,
            active              INTEGER DEFAULT 1,
            created_at          TEXT NOT NULL,
            require_scan_checkout INTEGER DEFAULT -1,
            parent_item_id      INTEGER REFERENCES items(id),
            ebay_status         TEXT DEFAULT 'not_listed',
            ebay_listing_id     TEXT,
            ebay_listed_price   REAL,
            ebay_listed_date    TEXT,
            company_id          INTEGER REFERENCES companies(id),
            po_number           TEXT,
            internal_sku        TEXT,
            resolution          TEXT,
            lens_type           TEXT,
            has_poe             INTEGER DEFAULT 0,
            wireless_standard   TEXT,
            port_count          INTEGER,
            poe_budget          TEXT,
            throughput          TEXT,
            os_type             TEXT,
            screen_size         TEXT,
            battery_life        TEXT,
            imei                TEXT,
            carrier             TEXT,
            cable_type          TEXT,
            cable_gauge         TEXT,
            connector_type      TEXT,
            cable_length        TEXT,
            extra_fields        TEXT DEFAULT '{}'
        );
        CREATE TABLE IF NOT EXISTS companies (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT UNIQUE NOT NULL,
            notes      TEXT,
            active     INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS category_fields (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id  INTEGER NOT NULL REFERENCES categories(id),
            field_label  TEXT NOT NULL,
            field_key    TEXT NOT NULL,
            field_type   TEXT NOT NULL DEFAULT 'text',
            placeholder  TEXT,
            required     INTEGER DEFAULT 0,
            sort_order   INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS product_catalog (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT NOT NULL,
            manufacturer    TEXT,
            model           TEXT,
            category_id     INTEGER REFERENCES categories(id),
            product_type_id INTEGER REFERENCES product_types(id),
            description     TEXT,
            default_cost    REAL,
            default_sale    REAL,
            image_url       TEXT,
            active          INTEGER DEFAULT 1,
            created_at      TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            ts           TEXT NOT NULL,
            action       TEXT NOT NULL,
            item_id      INTEGER,
            item_name    TEXT,
            detail       TEXT,
            before_state TEXT,
            after_state  TEXT,
            username     TEXT
        );
    """)
    # Migrate existing DBs — add new columns if they don't exist
    existing = [r[1] for r in db.execute("PRAGMA table_info(items)")]
    migrations = [
        ("manufacturer",          "TEXT"),
        ("condition",             "TEXT DEFAULT 'New'"),
        ("owner_company",         "TEXT"),
        ("purchase_date",         "TEXT"),
        ("low_stock_threshold",   "INTEGER DEFAULT 0"),
        ("cpu",                   "TEXT"),
        ("ram",                   "TEXT"),
        ("storage",               "TEXT"),
        ("tax_rate",              "REAL DEFAULT 0"),
        ("sold",                  "INTEGER DEFAULT 0"),
        ("sold_date",             "TEXT"),
        ("sold_price",            "REAL"),
        ("sold_to",               "TEXT"),
        ("shelf",                 "TEXT"),
        ("require_scan_checkout", "INTEGER DEFAULT -1"),
        ("parent_item_id",        "INTEGER"),
        ("ebay_status",           "TEXT DEFAULT 'not_listed'"),
        ("ebay_listing_id",       "TEXT"),
        ("ebay_listed_price",     "REAL"),
        ("ebay_listed_date",      "TEXT"),
        ("company_id",            "INTEGER"),
        ("po_number",             "TEXT"),
        ("internal_sku",          "TEXT"),
        ("resolution",            "TEXT"),
        ("lens_type",             "TEXT"),
        ("has_poe",               "INTEGER DEFAULT 0"),
        ("wireless_standard",     "TEXT"),
        ("port_count",            "INTEGER"),
        ("poe_budget",            "TEXT"),
        ("throughput",            "TEXT"),
        ("os_type",               "TEXT"),
        ("screen_size",           "TEXT"),
        ("battery_life",          "TEXT"),
        ("imei",                  "TEXT"),
        ("carrier",               "TEXT"),
        ("cable_type",            "TEXT"),
        ("cable_gauge",           "TEXT"),
        ("connector_type",        "TEXT"),
        ("cable_length",          "TEXT"),
        ("extra_fields",          "TEXT DEFAULT '{}'"),
    ]
    for col, typedef in migrations:
        if col not in existing:
            try:
                db.execute(f"ALTER TABLE items ADD COLUMN {col} {typedef}")
            except Exception:
                pass
    # Remove low_stock_threshold from product_types if it exists (moved to items)
    # Migrate users table
    pt_existing = [r[1] for r in db.execute("PRAGMA table_info(product_types)")]
    if "require_scan_checkout" not in pt_existing:
        db.execute("ALTER TABLE product_types ADD COLUMN require_scan_checkout INTEGER DEFAULT 0")

    u_existing = [r[1] for r in db.execute("PRAGMA table_info(users)")]
    if "permissions" not in u_existing:
        db.execute("ALTER TABLE users ADD COLUMN permissions TEXT NOT NULL DEFAULT ''")

    # Performance indexes
    db.executescript("""
        CREATE INDEX IF NOT EXISTS idx_items_active       ON items(active);
        CREATE INDEX IF NOT EXISTS idx_items_serial       ON items(serial);
        CREATE INDEX IF NOT EXISTS idx_items_category     ON items(category_id);
        CREATE INDEX IF NOT EXISTS idx_items_checked_out  ON items(checked_out);
        CREATE INDEX IF NOT EXISTS idx_items_sold         ON items(sold);
        CREATE INDEX IF NOT EXISTS idx_audit_ts           ON audit_log(ts DESC);
        CREATE INDEX IF NOT EXISTS idx_audit_item         ON audit_log(item_id);
        CREATE INDEX IF NOT EXISTS idx_audit_user         ON audit_log(username);
        CREATE INDEX IF NOT EXISTS idx_mods_item          ON item_modifications(item_id);
        CREATE INDEX IF NOT EXISTS idx_items_parent       ON items(parent_item_id);
        CREATE INDEX IF NOT EXISTS idx_items_ebay         ON items(ebay_status);
        CREATE INDEX IF NOT EXISTS idx_items_sku          ON items(internal_sku);
        CREATE INDEX IF NOT EXISTS idx_items_company      ON items(company_id);
        CREATE INDEX IF NOT EXISTS idx_catalog_cat        ON product_catalog(category_id);
    """)
    # Seed default companies if empty
    if not db.execute("SELECT COUNT(*) FROM companies").fetchone()[0]:
        for co in ["Amazon", "B&H Photo", "CDW", "Adorama", "Newegg", "Insight", "Dell Technologies", "Other"]:
            try: db.execute("INSERT INTO companies (name) VALUES (?)", [co])
            except: pass

    db.commit()

def log_action(action, item_id=None, item_name=None, detail=None, before=None, after=None):
    username = session.get("username", "system")
    execute("""INSERT INTO audit_log (ts,action,item_id,item_name,detail,before_state,after_state,username)
               VALUES (?,?,?,?,?,?,?,?)""",
            [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), action, item_id, item_name, detail,
             json.dumps(before) if before else None, json.dumps(after) if after else None, username])

def hash_pw(pw):
    import hashlib, hmac
    return hmac.new(b"storelax-salt", pw.encode(), hashlib.sha256).hexdigest()

# ── Permissions ────────────────────────────────────────────────────────────────
ALL_PERMISSIONS = [
    ("view_inventory",  "View Inventory",    "See the items list and search"),
    ("view_dashboard",  "View Dashboard",    "Access charts and financial summary"),
    ("view_audit",      "View Audit Log",    "See the full audit trail"),
    ("checkout_checkin","Checkout / Check In","Check items in and out of inventory"),
    ("write_items",     "Add / Edit Items",  "Create new items and edit existing ones"),
    ("qty_adjust",      "Adjust Quantities", "Add, remove or set stock quantities"),
    ("sell_items",      "Mark as Sold",      "Mark items as sold and record sale price"),
    ("delete_items",    "Delete Items",      "Permanently delete items"),
    ("import_export",   "Import / Export",   "Import Excel files and export CSV/Excel"),
]
PERM_KEYS = [p[0] for p in ALL_PERMISSIONS]

# Admin role always has all permissions
ADMIN_DEFAULT_PERMS = set(PERM_KEYS)
# Default worker permissions — full create/edit access but no destructive ops
WORKER_DEFAULT_PERMS = {"view_inventory", "view_dashboard", "view_audit", "checkout_checkin", "write_items", "qty_adjust", "sell_items"}

def get_user_perms(user_id=None, role=None, perm_str=None):
    """Returns set of permission keys for the user."""
    if role == "admin":
        return ADMIN_DEFAULT_PERMS
    # Empty string means no custom permissions set yet — use role defaults
    if not perm_str:
        return WORKER_DEFAULT_PERMS
    stored = set(perm_str.split(",")) if perm_str else set()
    # Only keep valid keys
    return stored & set(PERM_KEYS)

def has_perm(perm):
    """Check if current session user has a permission."""
    if session.get("role") == "admin":
        return True
    perms = set(session.get("permissions", "").split(","))
    return perm in perms

def perm_required(perm):
    """Decorator: require a specific permission."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get("user_id"):
                if request.path.startswith("/api/") or request.path.startswith("/export") or request.path.startswith("/import"):
                    return jsonify({"ok": False, "msg": "Not logged in"}), 401
                return redirect(url_for("login_page"))
            if not has_perm(perm):
                if request.path.startswith("/api/") or request.path.startswith("/export") or request.path.startswith("/import"):
                    return jsonify({"ok": False, "msg": f"Permission denied: {perm}"}), 403
                return redirect(url_for("inventory"))
            return f(*args, **kwargs)
        return decorated
    return decorator

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/") or request.path.startswith("/export") or request.path.startswith("/import"):
                return jsonify({"ok": False, "msg": "Not logged in"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            if request.path.startswith("/api/") or request.path.startswith("/export") or request.path.startswith("/import"):
                return jsonify({"ok": False, "msg": "Admin only"}), 403
            return redirect(url_for("inventory"))
        return f(*args, **kwargs)
    return decorated

def get_low_stock_alerts():
    rows = query("""
        SELECT i.id, i.name, i.qty, i.qty_out, i.low_stock_threshold
        FROM items i
        WHERE i.active=1 AND i.qty IS NOT NULL
          AND i.low_stock_threshold > 0
          AND (i.qty - COALESCE(i.qty_out,0)) <= i.low_stock_threshold
        ORDER BY (i.qty - COALESCE(i.qty_out,0)) ASC
    """)
    return [dict(r) for r in rows]

PRODUCT_TYPES = [
    # ── Compute ──────────────────────────────────────────────────────────────
    # (name, serial_required, sku_required, qty_tracked)
    # serial_required: 1=required, 0=optional, -1=hidden
    # qty_tracked: 1=track qty, 0=individual checkout
    ("Desktop PC",                  1,  0,  0),   # serial required, checkout per unit
    ("Laptop",                      1,  0,  0),
    ("Server",                      1,  0,  0),
    ("Workstation",                 1,  0,  0),
    ("Thin Client",                 1,  0,  0),
    # ── Networking — Switches ─────────────────────────────────────────────
    ("24-Port UniFi Switch",        1,  0,  0),
    ("48-Port UniFi Switch",        1,  0,  0),
    ("8-Port UniFi Switch",         1,  0,  0),
    ("UniFi Switch (Other)",        1,  0,  0),
    # ── Networking — Wireless ─────────────────────────────────────────────
    ("UniFi Access Point",          1,  0,  0),
    ("UniFi Dream Machine / UDM",   1,  0,  0),
    # ── Networking — Security / Firewall ──────────────────────────────────
    ("Meraki Firewall / MX",        1,  0,  0),
    ("Meraki Switch",               1,  0,  0),
    ("Meraki Access Point",         1,  0,  0),
    ("Firewall (Other)",            1,  0,  0),
    # ── Networking — General ──────────────────────────────────────────────
    ("Router",                      1,  0,  0),
    ("Network Module / SFP",        0,  1,  1),   # qty-tracked, SKU-based
    # ── Cabling & Patching ────────────────────────────────────────────────
    ("Patch Cable — Cat5e",         0, -1,  1),   # qty tracked, no serial/SKU needed
    ("Patch Cable — Cat6",          0, -1,  1),
    ("Patch Cable — Cat6A",         0, -1,  1),
    ("Fiber Patch Cable",           0, -1,  1),
    ("Patch Panel",                 1,  0,  0),
    ("Keystone Jack",               0, -1,  1),
    ("Bulk Cable (ft)",             0, -1,  1),
    # ── Power & Rackmount ─────────────────────────────────────────────────
    ("UPS Battery Backup",          1,  0,  0),
    ("PDU / Power Strip",           1,  0,  0),
    ("Rack / Enclosure",            1,  0,  0),
    ("KVM Switch",                  1,  0,  0),
    # ── Storage & Memory ──────────────────────────────────────────────────
    ("Hard Drive / SSD",            1,  1,  0),   # serial + SKU both useful
    ("RAM Module",                  0,  1,  1),
    ("NAS / Storage Array",         1,  0,  0),
    ("USB / Flash Drive",           0, -1,  1),
    # ── Peripherals & Misc ────────────────────────────────────────────────
    ("Monitor",                     1,  0,  0),
    ("Keyboard / Mouse",            0, -1,  1),
    ("Docking Station",             1,  0,  0),
    ("IP Camera / NVR",             1,  0,  0),
    ("VoIP Phone",                  1,  0,  0),
    ("Printer / MFP",               1,  0,  0),
    # ── Consumables & Tools ───────────────────────────────────────────────
    ("Toner / Ink Cartridge",       0,  1,  1),
    ("Tools / Crimpers / Testers",  0, -1,  1),
    ("Mounting Hardware",           0, -1,  1),
    ("Miscellaneous Consumable",    0, -1,  1),
    # ── General ───────────────────────────────────────────────────────────
    ("For Sale",                    0,  1,  0),
    ("Serialized Asset",            1,  0,  0),
    ("One-Off / Other",             0,  0,  0),
]

def seed_db():
    with app.app_context():
        init_db()
        if query("SELECT COUNT(*) FROM users", one=True)[0] > 0:
            return
        execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                ["admin", hash_pw("admin123"), "admin"])
        execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                ["worker", hash_pw("worker123"), "worker"])
        for name, sr, skur, qty_t in PRODUCT_TYPES:
            execute("INSERT INTO product_types (name,serial_required,sku_required,qty_tracked) VALUES (?,?,?,?)",
                    [name, sr, skur, qty_t])
        print("Fresh install ready | Logins: admin/admin123  and  worker/worker123")

# ── Auth routes ────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET","POST"])
def login_page():
    if session.get("user_id"):
        return redirect(url_for("inventory"))
    error = None
    if request.method == "POST":
        ip = request.remote_addr or "unknown"
        allowed, reset_in = check_login_rate(ip)
        if not allowed:
            return render_template("login.html", error=f"Too many login attempts. Try again in {reset_in} seconds."), 429
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        user = query("SELECT * FROM users WHERE username=?", [username], one=True)
        if user and user["password"] == hash_pw(password):
            clear_login_rate(ip)  # reset counter on success
            perms = get_user_perms(user["id"], user["role"], user["permissions"] if "permissions" in user.keys() else "")
            session.clear()  # prevent session fixation
            session["user_id"]     = user["id"]
            session["username"]    = user["username"]
            session["role"]        = user["role"]
            session["permissions"] = ",".join(perms)
            return redirect(url_for("inventory"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

# ── Pages ──────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def inventory():
    cats      = query("SELECT * FROM categories ORDER BY name")
    ptypes    = [dict(r) for r in query("SELECT * FROM product_types WHERE active=1 ORDER BY name")]
    companies = [dict(r) for r in query("SELECT * FROM companies WHERE active=1 ORDER BY name")]
    # Build category field map: {cat_id: [fields]}
    all_cat_fields = {}
    for row in query("SELECT * FROM category_fields ORDER BY category_id, sort_order, id"):
        cid = str(row["category_id"])
        all_cat_fields.setdefault(cid, []).append(dict(row))
    alerts    = get_low_stock_alerts()
    return render_template("inventory.html", categories=cats, product_types=ptypes,
                           companies=companies, category_fields=all_cat_fields,
                           alerts=alerts,
                           user_perms=list(session.get("permissions","").split(",")))

@app.route("/checked-out")
@login_required
def checked_out_page():
    cats = query("SELECT * FROM categories ORDER BY name")
    rows = query("""
        SELECT i.*, c.name as category, c.color, pt.name as product_type
        FROM items i
        LEFT JOIN categories c ON c.id=i.category_id
        LEFT JOIN product_types pt ON pt.id=i.product_type_id
        WHERE i.active=1 AND i.checked_out=1
        ORDER BY i.checkout_date DESC
    """)
    items = [dict(r) for r in rows]
    return render_template("checked_out.html", items=items, categories=cats)

@app.route("/audit")
@login_required
@perm_required("view_audit")
def audit():
    return render_template("audit.html")

@app.route("/dashboard")
@login_required
@perm_required("view_dashboard")
def dashboard():
    return render_template("dashboard.html", user_perms=list(session.get("permissions","").split(",")))

@app.route("/admin")
@login_required
@admin_required
def admin_page():
    users     = [dict(r) for r in query("SELECT id,username,role,permissions FROM users ORDER BY role,username")]
    for u in users:
        u["perm_set"] = set(u["permissions"].split(",")) if u["permissions"] else set()
    ptypes    = [dict(r) for r in query("SELECT * FROM product_types ORDER BY name")]
    companies = [dict(r) for r in query("SELECT * FROM companies WHERE active=1 ORDER BY name")]
    return render_template("admin.html", users=users, product_types=ptypes,
                           companies=companies,
                           all_permissions=ALL_PERMISSIONS, perm_keys=PERM_KEYS)

# ── API: Financial / Dashboard ─────────────────────────────────────────────────

@app.route("/api/dashboard")
@login_required
@perm_required("view_dashboard")
def api_dashboard():
    # Summary cards
    total      = query("SELECT COUNT(*) FROM items WHERE active=1 AND sold=0", one=True)[0]
    checked_out= query("SELECT COUNT(*) FROM items WHERE active=1 AND checked_out=1 AND sold=0", one=True)[0]
    low_stock  = query("SELECT COUNT(*) FROM items WHERE active=1 AND sold=0 AND qty IS NOT NULL AND low_stock_threshold>0 AND (qty-COALESCE(qty_out,0))<=low_stock_threshold", one=True)[0]
    sold_count = query("SELECT COUNT(*) FROM items WHERE active=1 AND sold=1", one=True)[0]

    # Stock value = cost * qty (or cost * 1 for serialized)
    val_row = query("""SELECT SUM(CASE WHEN qty IS NULL THEN COALESCE(cost_price,0)
                                       ELSE COALESCE(cost_price,0)*COALESCE(qty,0) END)
                       FROM items WHERE active=1 AND sold=0""", one=True)
    stock_value = round(val_row[0] or 0, 2)

    # Tax paid total
    tax_row = query("""SELECT SUM(CASE WHEN tax_paid=1 THEN
                           COALESCE(cost_price,0)*(COALESCE(tax_rate,8.875)/100)*
                           CASE WHEN qty IS NULL THEN 1 ELSE COALESCE(qty,0) END
                       ELSE 0 END) FROM items WHERE active=1""", one=True)
    total_tax = round(tax_row[0] or 0, 2)

    # Realized revenue from sold items
    rev_row = query("SELECT SUM(COALESCE(sold_price,sale_price,0)) FROM items WHERE active=1 AND sold=1", one=True)
    revenue  = round(rev_row[0] or 0, 2)
    cost_sold= query("SELECT SUM(COALESCE(cost_price,0)) FROM items WHERE active=1 AND sold=1", one=True)[0] or 0
    profit   = round(revenue - cost_sold, 2)

    # Stock value by category (for donut chart)
    cat_rows = query("""SELECT c.name, c.color,
                               SUM(CASE WHEN i.qty IS NULL THEN COALESCE(i.cost_price,0)
                                        ELSE COALESCE(i.cost_price,0)*COALESCE(i.qty,0) END) as val,
                               COUNT(*) as cnt
                        FROM items i
                        JOIN categories c ON c.id=i.category_id
                        WHERE i.active=1 AND i.sold=0
                        GROUP BY c.id ORDER BY val DESC""")
    by_category = [{"name":r["name"],"color":r["color"],"value":round(r["val"] or 0,2),"count":r["cnt"]} for r in cat_rows]

    # Most checked-out items (from audit log)
    top_rows = query("""SELECT item_name, COUNT(*) as cnt FROM audit_log
                        WHERE action='CHECKOUT' AND item_name IS NOT NULL
                        GROUP BY item_name ORDER BY cnt DESC LIMIT 8""")
    top_items = [{"name":r["item_name"],"count":r["cnt"]} for r in top_rows]

    # Tech activity
    tech_rows = query("""SELECT
                           CASE WHEN detail LIKE 'By: %' THEN
                               TRIM(SUBSTR(detail, 5, CASE WHEN INSTR(detail,' | ')>0 THEN INSTR(detail,' | ')-5 ELSE LENGTH(detail) END))
                           ELSE username END as tech,
                           COUNT(*) as cnt
                         FROM audit_log WHERE action='CHECKOUT'
                         GROUP BY tech ORDER BY cnt DESC LIMIT 6""")
    tech_activity = [{"name":r["tech"],"count":r["cnt"]} for r in tech_rows]

    # Monthly spend (last 6 months by purchase_date)
    monthly = query("""SELECT SUBSTR(purchase_date,1,7) as month,
                              SUM(COALESCE(cost_price,0)) as spend,
                              COUNT(*) as items
                       FROM items WHERE active=1 AND purchase_date IS NOT NULL AND purchase_date!=''
                       GROUP BY month ORDER BY month DESC LIMIT 6""")
    monthly_spend = [{"month":r["month"],"spend":round(r["spend"] or 0,2),"items":r["items"]} for r in monthly]

    # Recent activity
    recent = query("""SELECT ts, action, item_name, username, detail FROM audit_log
                      ORDER BY id DESC LIMIT 12""")
    activity = [dict(r) for r in recent]

    # Low stock items detail
    low_items = get_low_stock_alerts()

    return jsonify({
        "cards": {"total":total,"checked_out":checked_out,"low_stock":low_stock,
                  "stock_value":stock_value,"total_tax":total_tax,
                  "revenue":revenue,"profit":profit,"sold_count":sold_count},
        "by_category": by_category,
        "top_items": top_items,
        "tech_activity": tech_activity,
        "monthly_spend": list(reversed(monthly_spend)),
        "activity": activity,
        "low_items": low_items,
    })

@app.route("/api/item/sell", methods=["POST"])
@login_required
@perm_required("sell_items")
def api_item_sell():
    d    = request.json
    item = query("SELECT * FROM items WHERE id=? AND active=1", [d["id"]], one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    if item["checked_out"]: return jsonify({"ok":False,"msg":"Item is currently checked out"})
    price = float(d.get("price") or item["sale_price"] or 0)
    sold_to = d.get("sold_to","").strip()
    now = datetime.now().strftime("%Y-%m-%d")
    execute("UPDATE items SET sold=1, sold_date=?, sold_price=?, sold_to=?, checked_out=0 WHERE id=?",
            [now, price, sold_to, item["id"]])
    profit = round(price - (item["cost_price"] or 0), 2)
    log_action("ITEM_SOLD", item["id"], item["name"],
               f"Sold to: {sold_to or 'unknown'} | Price: ${price:.2f} | Profit: ${profit:.2f}",
               {"sold":0}, {"sold":1,"price":price})
    return jsonify({"ok":True,"profit":profit})

# ── API: Items ─────────────────────────────────────────────────────────────────

@app.route("/api/items")
@login_required
def api_items():
    search = request.args.get("q","").strip()
    cat_id = request.args.get("cat","")
    status = request.args.get("status","")
    sort   = request.args.get("sort","name")
    pt_id  = request.args.get("ptype","")
    sql = """SELECT i.*, c.name as category, c.color, pt.name as product_type, pt.qty_tracked,
                    co.name as company_name
             FROM items i
             LEFT JOIN categories c ON c.id=i.category_id
             LEFT JOIN product_types pt ON pt.id=i.product_type_id
             LEFT JOIN companies co ON co.id=i.company_id
             WHERE i.active=1"""
    args = []
    if search:
        sql += " AND (i.name LIKE ? OR i.serial LIKE ? OR i.model LIKE ? OR i.sku LIKE ? OR i.internal_sku LIKE ? OR CAST(i.shelf AS TEXT) LIKE ? OR i.job_ref LIKE ? OR i.owner_company LIKE ? OR i.manufacturer LIKE ? OR i.po_number LIKE ?)"
        s = f"%{search}%"; args += [s,s,s,s,s,s,s,s,s,s]
    if cat_id: sql += " AND i.category_id=?"; args.append(cat_id)
    if pt_id:  sql += " AND i.product_type_id=?"; args.append(pt_id)
    if status == "out":    sql += " AND i.checked_out=1"
    elif status == "in":   sql += " AND i.checked_out=0"
    elif status == "low":  sql += " AND i.qty IS NOT NULL AND i.low_stock_threshold > 0 AND (i.qty - COALESCE(i.qty_out,0)) <= i.low_stock_threshold"
    order = {"name":"i.name","shelf":"i.shelf","cat":"c.name","cost":"i.cost_price","sale":"i.sale_price","date":"i.purchase_date"}.get(sort,"i.name")
    sql += f" ORDER BY {order}"
    result = []
    for r in query(sql, args):
        d = dict(r)
        d["available"] = (d["qty"] or 0) - (d["qty_out"] or 0) if d["qty"] is not None else None
        d["is_low"] = bool(d["low_stock_threshold"] and d["available"] is not None and d["available"] <= d["low_stock_threshold"])
        d["profit"] = round(d["sale_price"] - d["cost_price"], 2) if d["sale_price"] and d["cost_price"] else None
        d["tax_amount"] = round(d["cost_price"] * (d["tax_rate"] or 0) / 100, 2) if d["cost_price"] and d["tax_paid"] == 1 else 0
        result.append(d)
    return jsonify(result)

@app.route("/api/scan")
@login_required
def api_scan():
    code = request.args.get("code","").strip()
    if not code: return jsonify({"found":False})
    row = query("""SELECT i.*, c.name as category, c.color, pt.name as product_type, pt.qty_tracked
                   FROM items i
                   LEFT JOIN categories c ON c.id=i.category_id
                   LEFT JOIN product_types pt ON pt.id=i.product_type_id
                   WHERE i.active=1 AND (i.serial=? OR i.model=? OR i.sku=?) LIMIT 1""",
                [code,code,code], one=True)
    return jsonify({"found":bool(row),"item":dict(row) if row else None})

@app.route("/api/checkout", methods=["POST"])
@login_required
@perm_required("checkout_checkin")
def api_checkout():
    d    = request.json
    item = query(
        "SELECT i.*, pt.require_scan_checkout as pt_require_scan "
        "FROM items i LEFT JOIN product_types pt ON pt.id=i.product_type_id "
        "WHERE i.id=? AND i.active=1", [d["id"]], one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    if item["sold"]: return jsonify({"ok":False,"msg":"Item has been sold"})
    if item["checked_out"]: return jsonify({"ok":False,"msg":"Already checked out"})

    # Resolve scan requirement: item override (-1=inherit) → product type default → 0
    item_override = item["require_scan_checkout"] if item["require_scan_checkout"] is not None else -1
    type_default  = item["pt_require_scan"] or 0
    scan_required = item_override if item_override != -1 else type_default

    if scan_required:
        scanned = d.get("scanned_code","").strip()
        if not scanned:
            return jsonify({"ok":False,"msg":"scan_required",
                            "detail":"This item requires a serial/SKU scan before checkout"})
        valid = [x for x in [item["serial"], item["sku"], item["model"]] if x]
        if scanned not in valid:
            return jsonify({"ok":False,"msg":"scan_mismatch",
                            "detail":f"Scanned '{scanned}' does not match item serial/SKU/model"})

    now = datetime.now().strftime("%m/%d/%y")
    who = d.get("who","").strip() or session.get("username","?")
    execute("UPDATE items SET checked_out=1,checkout_date=?,checkout_by=?,job_ref=? WHERE id=?",
            [now, who, d.get("job_ref",""), item["id"]])
    log_action("CHECKOUT",item["id"],item["name"],f"By: {who} | Job: {d.get('job_ref','')}",
               {"checked_out":0},{"checked_out":1})
    return jsonify({"ok":True})

@app.route("/api/checkin", methods=["POST"])
@login_required
@perm_required("checkout_checkin")
def api_checkin():
    d    = request.json
    item = query("SELECT * FROM items WHERE id=? AND active=1",[d["id"]],one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    if not item["checked_out"]: return jsonify({"ok":False,"msg":"Already checked in"})
    execute("UPDATE items SET checked_out=0,checkout_date=NULL,checkout_by=NULL,job_ref=NULL WHERE id=?",
            [item["id"]])
    log_action("CHECKIN",item["id"],item["name"],f"Returned. Was on: {item['job_ref'] or '-'}",
               {"checked_out":1},{"checked_out":0})
    return jsonify({"ok":True})

@app.route("/api/qty_adjust", methods=["POST"])
@login_required
@perm_required("qty_adjust")
def api_qty_adjust():
    d    = request.json
    item = query("SELECT * FROM items WHERE id=? AND active=1",[d["id"]],one=True)
    if not item or item["qty"] is None: return jsonify({"ok":False,"msg":"Not qty-tracked"})
    before = {"qty":item["qty"],"qty_out":item["qty_out"]}
    amount = int(d.get("amount",0)); note = d.get("note",""); action = d.get("action")
    if action=="add":
        nq = item["qty"]+amount; execute("UPDATE items SET qty=? WHERE id=?",[nq,item["id"]])
        log_action("QTY_ADD",item["id"],item["name"],f"+{amount}. {note}. Total:{nq}",before,{"qty":nq})
    elif action=="remove":
        no = (item["qty_out"] or 0)+amount; execute("UPDATE items SET qty_out=? WHERE id=?",[no,item["id"]])
        log_action("QTY_REMOVE",item["id"],item["name"],f"-{amount}. {note}. Left:{item['qty']-no}",before,{"qty_out":no})
    elif action=="set":
        execute("UPDATE items SET qty=?,qty_out=0 WHERE id=?",[amount,item["id"]])
        log_action("QTY_SET",item["id"],item["name"],f"Set to {amount}. {note}",before,{"qty":amount})
    u = query("SELECT qty,qty_out FROM items WHERE id=?",[item["id"]],one=True)
    return jsonify({"ok":True,"qty":u["qty"],"qty_out":u["qty_out"]})

def _save_item(d, iid=None):
    """Shared save logic for add and edit."""
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # On edit, preserve existing qty if the field wasn't shown (non-qty-tracked types)
    if iid and d.get("qty") in (None, ""):
        existing = query("SELECT qty FROM items WHERE id=?", [iid], one=True)
        qty = existing["qty"] if existing else None
    else:
        qty = int(d["qty"]) if d.get("qty") not in (None,"") else None
    cost = float(d["cost_price"]) if d.get("cost_price") not in (None,"") else None
    sale = float(d["sale_price"]) if d.get("sale_price") not in (None,"") else None
    tax  = int(d.get("tax_paid", -1))   # -1=not set, 0=no, 1=yes
    rate = float(d.get("tax_rate",8.875)) if d.get("tax_rate") not in (None,"") else 8.875
    # Preserve existing threshold on edit if not explicitly provided
    if iid and d.get("low_stock_threshold") in (None, ""):
        existing_lst = query("SELECT low_stock_threshold FROM items WHERE id=?", [iid], one=True)
        lst = existing_lst["low_stock_threshold"] if existing_lst else 0
    else:
        lst = int(d.get("low_stock_threshold", 0)) if d.get("low_stock_threshold") not in (None, "") else 0
    fields = dict(
        name=d.get("name"), manufacturer=d.get("manufacturer") or None,
        model=d.get("model") or None, serial=d.get("serial") or None,
        sku=d.get("sku") or None, category_id=d.get("category_id") or None,
        product_type_id=d.get("product_type_id") or None,
        condition=d.get("condition") or "New",
        owner_company=d.get("owner_company") or None,
        purchase_date=d.get("purchase_date") or None,
        shelf=d.get("shelf") or None, qty=qty,
        low_stock_threshold=lst, notes=d.get("notes",""),
        cpu=d.get("cpu") or None, ram=d.get("ram") or None,
        storage=d.get("storage") or None,
        cost_price=cost, sale_price=sale, tax_paid=tax, tax_rate=rate,
        sold=int(d.get("sold",0)), sold_date=d.get("sold_date") or None,
        sold_price=float(d["sold_price"]) if d.get("sold_price") not in (None,"") else None,
        sold_to=d.get("sold_to") or None,
        require_scan_checkout=int(d["require_scan_checkout"]) if d.get("require_scan_checkout") not in (None,"") else -1,
        ebay_status=d.get("ebay_status") or "not_listed",
        ebay_listing_id=d.get("ebay_listing_id") or None,
        ebay_listed_price=float(d["ebay_listed_price"]) if d.get("ebay_listed_price") not in (None,"") else None,
        ebay_listed_date=d.get("ebay_listed_date") or None,
        company_id=int(d["company_id"]) if d.get("company_id") not in (None,"") else None,
        po_number=d.get("po_number") or None,
        # category-specific fields
        resolution=d.get("resolution") or None,
        lens_type=d.get("lens_type") or None,
        has_poe=int(d.get("has_poe",0)),
        wireless_standard=d.get("wireless_standard") or None,
        port_count=int(d["port_count"]) if d.get("port_count") not in (None,"") else None,
        poe_budget=d.get("poe_budget") or None,
        throughput=d.get("throughput") or None,
        os_type=d.get("os_type") or None,
        screen_size=d.get("screen_size") or None,
        battery_life=d.get("battery_life") or None,
        imei=d.get("imei") or None,
        carrier=d.get("carrier") or None,
        cable_type=d.get("cable_type") or None,
        cable_gauge=d.get("cable_gauge") or None,
        connector_type=d.get("connector_type") or None,
        cable_length=d.get("cable_length") or None,
        extra_fields=json.dumps(d.get("extra_fields") or {}),
    )
    if iid:
        cols = ", ".join(f"{k}=?" for k in fields)
        execute(f"UPDATE items SET {cols} WHERE id=?", list(fields.values()) + [iid])
        # Auto-generate internal_sku if item still has none and no vendor sku
        row = query("SELECT sku, internal_sku, category_id FROM items WHERE id=?", [iid], one=True)
        if row and not row["sku"] and not row["internal_sku"]:
            _ensure_internal_sku(iid, row["category_id"])
        return iid
    else:
        fields["created_at"] = now
        cols = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)
        new_id = execute(f"INSERT INTO items ({cols}) VALUES ({placeholders})", list(fields.values()))
        # Auto-generate internal SKU if no vendor SKU provided
        if not fields.get("sku"):
            _ensure_internal_sku(new_id, fields.get("category_id"))
        return new_id

def _ensure_internal_sku(item_id, category_id):
    """Generate and assign CAT-YYYYMMDD-NNN internal SKU if not already set."""
    from datetime import date as _date
    cat = query("SELECT name FROM categories WHERE id=?", [category_id], one=True) if category_id else None
    cat_code = (cat["name"][:4].upper().replace(" ","") if cat else "GEN")
    today = _date.today().strftime("%Y%m%d")
    prefix = f"{cat_code}-{today}-"
    # Find highest seq for this cat+date prefix
    existing = query(
        "SELECT internal_sku FROM items WHERE internal_sku LIKE ?",
        [f"{prefix}%"])
    nums = []
    for r in existing:
        try: nums.append(int(r["internal_sku"].split("-")[-1]))
        except: pass
    seq = (max(nums) + 1) if nums else 1
    sku = f"{prefix}{seq:03d}"
    execute("UPDATE items SET internal_sku=? WHERE id=?", [sku, item_id])

@app.route("/api/item/add", methods=["POST"])
@login_required
@perm_required("write_items")
def api_item_add():
    d = request.json
    if not d.get("name","").strip():
        return jsonify({"ok":False,"msg":"Name required"})
    # Validate foreign keys exist
    cat_id = d.get("category_id")
    if cat_id and not query("SELECT id FROM categories WHERE id=?", [cat_id], one=True):
        return jsonify({"ok":False,"msg":"Invalid category"})
    pt_id = d.get("product_type_id")
    if pt_id and not query("SELECT id FROM product_types WHERE id=?", [pt_id], one=True):
        return jsonify({"ok":False,"msg":"Invalid product type"})
    try:
        iid = _save_item(d)
        log_action("ITEM_ADD", iid, d.get("name"), f"Shelf {d.get('shelf')}")
        return jsonify({"ok":True,"id":iid})
    except Exception as e:
        return jsonify({"ok":False,"msg":f"Save failed: {str(e)}"})

@app.route("/api/item/edit", methods=["POST"])
@login_required
@perm_required("write_items")
def api_item_edit():
    d    = request.json
    if not d.get("id"): return jsonify({"ok":False,"msg":"ID required"})
    item = query("SELECT * FROM items WHERE id=? AND active=1", [d["id"]], one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    if not d.get("name","").strip():
        return jsonify({"ok":False,"msg":"Name required"})
    cat_id = d.get("category_id")
    if cat_id and not query("SELECT id FROM categories WHERE id=?", [cat_id], one=True):
        return jsonify({"ok":False,"msg":"Invalid category"})
    try:
        _save_item(d, d["id"])
        log_action("ITEM_EDIT", d["id"], d.get("name"), "Edited", dict(item), d)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"msg":f"Save failed: {str(e)}"})

@app.route("/api/item/delete", methods=["POST"])
@login_required
@perm_required("delete_items")
def api_item_delete():
    d   = request.json
    ids = d.get("ids") or ([d["id"]] if d.get("id") else [])
    for iid in ids:
        item = query("SELECT * FROM items WHERE id=?",[iid],one=True)
        if item:
            execute("UPDATE items SET active=0 WHERE id=?",[iid])
            log_action("ITEM_DELETE",iid,item["name"],"Deleted")
    return jsonify({"ok":True,"deleted":len(ids)})

@app.route("/api/category/add", methods=["POST"])
@login_required
def api_cat_add():
    d = request.json
    try:
        cid = execute("INSERT INTO categories (name,color) VALUES (?,?)",[d["name"],d.get("color","#1d4ed8")])
        log_action("CAT_ADD",detail=f"Added: {d['name']}")
        return jsonify({"ok":True,"id":cid,"name":d["name"],"color":d.get("color","#1d4ed8")})
    except: return jsonify({"ok":False,"msg":"Already exists"})

@app.route("/api/product_types")
@login_required
def api_product_types():
    return jsonify([dict(r) for r in query("SELECT * FROM product_types WHERE active=1 ORDER BY name")])

@app.route("/api/product_type/add", methods=["POST"])
@login_required
@admin_required
def api_pt_add():
    d = request.json
    if not d.get("name"): return jsonify({"ok":False,"msg":"Name required"})
    try:
        pid = execute(
            "INSERT INTO product_types (name,serial_required,sku_required,qty_tracked,require_scan_checkout) VALUES (?,?,?,?,?)",
            [d["name"],int(d.get("serial_required",0)),int(d.get("sku_required",0)),
             int(d.get("qty_tracked",0)),int(d.get("require_scan_checkout",0))])
        return jsonify({"ok":True,"id":pid})
    except: return jsonify({"ok":False,"msg":"Already exists"})

@app.route("/api/product_type/edit", methods=["POST"])
@login_required
@admin_required
def api_pt_edit():
    d = request.json
    execute("UPDATE product_types SET name=?,serial_required=?,sku_required=?,qty_tracked=?,require_scan_checkout=? WHERE id=?",
            [d["name"],int(d.get("serial_required",0)),int(d.get("sku_required",0)),
             int(d.get("qty_tracked",0)),int(d.get("require_scan_checkout",0)),d["id"]])
    return jsonify({"ok":True})

# ── Item modifications ─────────────────────────────────────────────────────────

@app.route("/api/item/modifications", methods=["GET"])
@login_required
def api_item_mods_get():
    iid = request.args.get("item_id")
    if not iid: return jsonify({"ok":False,"msg":"item_id required"})
    mods = query("""
        SELECT m.*, i.name as spawned_name
        FROM item_modifications m
        LEFT JOIN items i ON i.id = m.spawned_item_id
        WHERE m.item_id = ?
        ORDER BY m.id DESC""", [iid])
    return jsonify([dict(r) for r in mods])

@app.route("/api/item/modify", methods=["POST"])
@login_required
@perm_required("write_items")
def api_item_modify():
    d = request.json
    iid = d.get("item_id")
    if not iid: return jsonify({"ok":False,"msg":"item_id required"})
    item = query("SELECT * FROM items WHERE id=? AND active=1", [iid], one=True)
    if not item: return jsonify({"ok":False,"msg":"Item not found"})
    field   = d.get("field_changed","").strip()
    old_val = d.get("old_value","")
    new_val = d.get("new_value","")
    notes   = d.get("notes","")
    if not field: return jsonify({"ok":False,"msg":"field_changed required"})
    MUTABLE_FIELDS = {"ram","cpu","storage","notes","condition","shelf","sale_price","cost_price"}
    if field in MUTABLE_FIELDS:
        try:
            execute(f"UPDATE items SET {field}=? WHERE id=?", [new_val or None, iid])
        except Exception as e:
            return jsonify({"ok":False,"msg":f"Field update failed: {e}"})
    spawned_id = None
    spawn_data = d.get("spawn_item")
    if spawn_data and spawn_data.get("name","").strip():
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        spawned_id = execute(
            "INSERT INTO items (name,category_id,product_type_id,condition,shelf,cost_price,notes,"
            "parent_item_id,created_at,active) VALUES (?,?,?,?,?,?,?,?,?,1)",
            [spawn_data["name"].strip(),
             spawn_data.get("category_id") or item["category_id"],
             spawn_data.get("product_type_id") or None,
             spawn_data.get("condition","Used"),
             spawn_data.get("shelf") or item["shelf"],
             float(spawn_data["cost_price"]) if spawn_data.get("cost_price") not in (None,"") else None,
             spawn_data.get("notes","") or f"Removed from: {item['name']} (ID {iid})",
             iid, now])
        log_action("ITEM_ADD", spawned_id, spawn_data["name"],
                   f"Spawned from modification of '{item['name']}' (ID {iid})")
    ts_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    execute(
        "INSERT INTO item_modifications (item_id,ts,modified_by,field_changed,old_value,new_value,notes,spawned_item_id) "
        "VALUES (?,?,?,?,?,?,?,?)",
        [iid, ts_now, session.get("username","?"), field, old_val, new_val, notes, spawned_id])
    log_action("ITEM_MODIFY", iid, item["name"],
               f"{field}: '{old_val}' -> '{new_val}'" + (f" | Spawned item ID {spawned_id}" if spawned_id else ""),
               {field: old_val}, {field: new_val})
    updated = query("SELECT * FROM items WHERE id=?", [iid], one=True)
    return jsonify({"ok":True,"spawned_id":spawned_id,"item":dict(updated)})

@app.route("/api/item/lineage")
@login_required
def api_item_lineage():
    iid = int(request.args.get("item_id",0))
    if not iid: return jsonify({"ok":False,"msg":"item_id required"})
    item = query("SELECT * FROM items WHERE id=?", [iid], one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    parent = None
    if item["parent_item_id"]:
        p = query("SELECT id,name,ram,cpu,storage,condition,shelf FROM items WHERE id=?",
                  [item["parent_item_id"]], one=True)
        if p: parent = dict(p)
    children = query(
        "SELECT id,name,ram,cpu,storage,condition,shelf,created_at FROM items WHERE parent_item_id=? AND active=1",
        [iid])
    return jsonify({"ok":True,"parent":parent,"children":[dict(c) for c in children]})

# ── eBay tracking ──────────────────────────────────────────────────────────────

@app.route("/api/item/ebay", methods=["POST"])
@login_required
@perm_required("sell_items")
def api_item_ebay():
    d = request.json
    iid = d.get("item_id")
    if not iid: return jsonify({"ok":False,"msg":"item_id required"})
    item = query("SELECT * FROM items WHERE id=? AND active=1", [iid], one=True)
    if not item: return jsonify({"ok":False,"msg":"Not found"})
    status       = d.get("ebay_status","not_listed")
    listing_id   = d.get("ebay_listing_id","").strip() or None
    listed_price = float(d["ebay_listed_price"]) if d.get("ebay_listed_price") not in (None,"") else None
    listed_date  = d.get("ebay_listed_date") or None
    if status not in ("not_listed","listed","sold"):
        return jsonify({"ok":False,"msg":"Invalid status. Use: not_listed, listed, sold"})
    execute(
        "UPDATE items SET ebay_status=?,ebay_listing_id=?,ebay_listed_price=?,ebay_listed_date=? WHERE id=?",
        [status, listing_id, listed_price, listed_date, iid])
    old_status = item["ebay_status"] or "not_listed"
    log_action("EBAY_UPDATE", iid, item["name"],
               f"eBay: {old_status} -> {status}" + (f" | Listing: {listing_id}" if listing_id else ""),
               {"ebay_status": old_status}, {"ebay_status": status})
    return jsonify({"ok":True})

@app.route("/api/product_type/delete", methods=["POST"])
@login_required
@admin_required
def api_pt_delete():
    execute("UPDATE product_types SET active=0 WHERE id=?",[request.json["id"]])
    return jsonify({"ok":True})

# ── Companies ──────────────────────────────────────────────────────────────────

@app.route("/api/companies")
@login_required
def api_companies():
    return jsonify([dict(r) for r in query("SELECT * FROM companies WHERE active=1 ORDER BY name")])

@app.route("/api/company/add", methods=["POST"])
@login_required
def api_company_add():
    d = request.json
    name = (d.get("name") or "").strip()
    if not name: return jsonify({"ok":False,"msg":"Name required"})
    try:
        cid = execute("INSERT INTO companies (name,notes) VALUES (?,?)", [name, d.get("notes","")])
        log_action("COMPANY_ADD", detail=f"Added company: {name}")
        return jsonify({"ok":True,"id":cid,"name":name})
    except: return jsonify({"ok":False,"msg":"Company already exists"})

@app.route("/api/company/edit", methods=["POST"])
@login_required
@admin_required
def api_company_edit():
    d = request.json
    execute("UPDATE companies SET name=?,notes=? WHERE id=?",
            [d["name"],d.get("notes",""),d["id"]])
    return jsonify({"ok":True})

@app.route("/api/company/delete", methods=["POST"])
@login_required
@admin_required
def api_company_delete():
    execute("UPDATE companies SET active=0 WHERE id=?", [request.json["id"]])
    return jsonify({"ok":True})

# ── Category Fields ────────────────────────────────────────────────────────────

@app.route("/api/category/fields")
@login_required
def api_category_fields():
    cat_id = request.args.get("cat_id")
    if cat_id:
        rows = query("SELECT * FROM category_fields WHERE category_id=? ORDER BY sort_order, id", [cat_id])
    else:
        rows = query("SELECT * FROM category_fields ORDER BY category_id, sort_order, id")
    return jsonify([dict(r) for r in rows])

@app.route("/api/category/fields/save", methods=["POST"])
@login_required
@admin_required
def api_category_fields_save():
    """Replace all field definitions for a category."""
    d = request.json
    cat_id = d.get("category_id")
    fields = d.get("fields", [])
    if not cat_id: return jsonify({"ok":False,"msg":"category_id required"})
    execute("DELETE FROM category_fields WHERE category_id=?", [cat_id])
    for i, f in enumerate(fields):
        label = (f.get("field_label") or "").strip()
        key   = (f.get("field_key") or "").strip().lower().replace(" ","_")
        if not label or not key: continue
        execute(
            "INSERT INTO category_fields (category_id,field_label,field_key,field_type,placeholder,required,sort_order) VALUES (?,?,?,?,?,?,?)",
            [cat_id, label, key, f.get("field_type","text"),
             f.get("placeholder",""), int(f.get("required",0)), i])
    return jsonify({"ok":True})

# ── Product Catalog ────────────────────────────────────────────────────────────

@app.route("/products")
@login_required
def products_page():
    cats   = query("SELECT * FROM categories ORDER BY name")
    ptypes = [dict(r) for r in query("SELECT * FROM product_types WHERE active=1 ORDER BY name")]
    all_cat_fields = {}
    for row in query("SELECT * FROM category_fields ORDER BY category_id, sort_order, id"):
        cid = str(row["category_id"])
        all_cat_fields.setdefault(cid, []).append(dict(row))
    return render_template("products.html", categories=cats, product_types=ptypes,
                           category_fields=all_cat_fields)

@app.route("/api/catalog")
@login_required
def api_catalog():
    q   = request.args.get("q","")
    cat = request.args.get("cat","")
    sql = """SELECT pc.*, c.name as category, pt.name as product_type
             FROM product_catalog pc
             LEFT JOIN categories c ON c.id=pc.category_id
             LEFT JOIN product_types pt ON pt.id=pc.product_type_id
             WHERE pc.active=1"""
    args = []
    if q:
        sql += " AND (pc.name LIKE ? OR pc.manufacturer LIKE ? OR pc.model LIKE ?)"
        s = f"%{q}%"; args += [s,s,s]
    if cat:
        sql += " AND pc.category_id=?"; args.append(cat)
    sql += " ORDER BY pc.name"
    return jsonify([dict(r) for r in query(sql, args)])

@app.route("/api/catalog/add", methods=["POST"])
@login_required
@perm_required("write_items")
def api_catalog_add():
    d = request.json
    if not d.get("name","").strip(): return jsonify({"ok":False,"msg":"Name required"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        cid = execute(
            "INSERT INTO product_catalog (name,manufacturer,model,category_id,product_type_id,"
            "description,default_cost,default_sale,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
            [d["name"].strip(), d.get("manufacturer") or None, d.get("model") or None,
             d.get("category_id") or None, d.get("product_type_id") or None,
             d.get("description") or None,
             float(d["default_cost"]) if d.get("default_cost") not in (None,"") else None,
             float(d["default_sale"]) if d.get("default_sale") not in (None,"") else None,
             now])
        return jsonify({"ok":True,"id":cid})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})

@app.route("/api/catalog/edit", methods=["POST"])
@login_required
@perm_required("write_items")
def api_catalog_edit():
    d = request.json
    execute("UPDATE product_catalog SET name=?,manufacturer=?,model=?,category_id=?,"
            "product_type_id=?,description=?,default_cost=?,default_sale=? WHERE id=?",
            [d["name"], d.get("manufacturer") or None, d.get("model") or None,
             d.get("category_id") or None, d.get("product_type_id") or None,
             d.get("description") or None,
             float(d["default_cost"]) if d.get("default_cost") not in (None,"") else None,
             float(d["default_sale"]) if d.get("default_sale") not in (None,"") else None,
             d["id"]])
    return jsonify({"ok":True})

@app.route("/api/catalog/delete", methods=["POST"])
@login_required
@perm_required("delete_items")
def api_catalog_delete():
    execute("UPDATE product_catalog SET active=0 WHERE id=?", [request.json["id"]])
    return jsonify({"ok":True})

@app.route("/api/catalog/to_item", methods=["POST"])
@login_required
@perm_required("write_items")
def api_catalog_to_item():
    """Create a new inventory item pre-filled from a catalog product."""
    d    = request.json
    prod = query("SELECT * FROM product_catalog WHERE id=? AND active=1", [d["id"]], one=True)
    if not prod: return jsonify({"ok":False,"msg":"Product not found"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    item_id = execute(
        "INSERT INTO items (name,manufacturer,model,category_id,product_type_id,"
        "cost_price,sale_price,condition,active,created_at) VALUES (?,?,?,?,?,?,?,?,1,?)",
        [prod["name"], prod["manufacturer"], prod["model"],
         prod["category_id"], prod["product_type_id"],
         prod["default_cost"], prod["default_sale"], "New", now])
    _ensure_internal_sku(item_id, prod["category_id"])
    log_action("ITEM_ADD", item_id, prod["name"], f"Created from catalog product ID {d['id']}")
    return jsonify({"ok":True,"item_id":item_id})

@app.route("/api/alerts")
@login_required
def api_alerts():
    return jsonify(get_low_stock_alerts())

@app.route("/api/audit")
@login_required
@perm_required("view_audit")
def api_audit():
    page=int(request.args.get("page",1)); limit=50; offset=(page-1)*limit
    search=request.args.get("q","")
    sql="SELECT * FROM audit_log"; args=[]
    if search:
        sql+=" WHERE action LIKE ? OR item_name LIKE ? OR detail LIKE ? OR username LIKE ?"
        s=f"%{search}%"; args=[s,s,s,s]
    total=query(f"SELECT COUNT(*) FROM ({sql})",args,one=True)[0]
    sql+=f" ORDER BY id DESC LIMIT {limit} OFFSET {offset}"
    return jsonify({"rows":[dict(r) for r in query(sql,args)],"total":total,"page":page})

# ── Admin: users ───────────────────────────────────────────────────────────────

@app.route("/api/user/add", methods=["POST"])
@login_required
@admin_required
def api_user_add():
    d = request.json
    if not d.get("username") or not d.get("password"):
        return jsonify({"ok":False,"msg":"Username and password required"})
    role = d.get("role","worker")
    # Build permissions string — admin gets all, worker gets defaults, custom overrides
    if role == "admin":
        perm_str = ",".join(ADMIN_DEFAULT_PERMS)
    else:
        custom = d.get("permissions")
        if custom is not None:
            perm_str = ",".join(set(custom) & set(PERM_KEYS))
        else:
            perm_str = ",".join(WORKER_DEFAULT_PERMS)
    try:
        uid = execute("INSERT INTO users (username,password,role,permissions) VALUES (?,?,?,?)",
                      [d["username"], hash_pw(d["password"]), role, perm_str])
        log_action("USER_ADD", detail=f"Added user: {d['username']} | role: {role}")
        return jsonify({"ok":True,"id":uid})
    except: return jsonify({"ok":False,"msg":"Username already exists"})

@app.route("/api/user/permissions", methods=["POST"])
@login_required
@admin_required
def api_user_permissions():
    d    = request.json
    uid  = d.get("id")
    perms = d.get("permissions", [])
    user = query("SELECT * FROM users WHERE id=?", [uid], one=True)
    if not user: return jsonify({"ok":False,"msg":"User not found"})
    if user["role"] == "admin":
        return jsonify({"ok":False,"msg":"Admin always has full permissions"})
    perm_str = ",".join(set(perms) & set(PERM_KEYS))
    execute("UPDATE users SET permissions=? WHERE id=?", [perm_str, uid])
    log_action("USER_PERMS", detail=f"Updated perms for {user['username']}: {perm_str}")
    return jsonify({"ok":True})

@app.route("/api/user/password", methods=["POST"])
@login_required
@admin_required
def api_user_password():
    d = request.json
    if not d.get("password"): return jsonify({"ok":False,"msg":"Password required"})
    execute("UPDATE users SET password=? WHERE id=?",[hash_pw(d["password"]),d["id"]])
    return jsonify({"ok":True})

@app.route("/api/user/delete", methods=["POST"])
@login_required
@admin_required
def api_user_delete():
    d = request.json
    if d["id"] == session.get("user_id"):
        return jsonify({"ok":False,"msg":"Cannot delete yourself"})
    execute("DELETE FROM users WHERE id=?",[d["id"]])
    return jsonify({"ok":True})

@app.route("/api/change_password", methods=["POST"])
@login_required
def api_change_password():
    d    = request.json
    user = query("SELECT * FROM users WHERE id=?",[session["user_id"]],one=True)
    if not user or user["password"] != hash_pw(d.get("old_password","")):
        return jsonify({"ok":False,"msg":"Current password incorrect"})
    execute("UPDATE users SET password=? WHERE id=?",[hash_pw(d["new_password"]),session["user_id"]])
    return jsonify({"ok":True})

# ── Export / Import ────────────────────────────────────────────────────────────

@app.route("/export/csv")
@login_required
@perm_required("import_export")
def export_csv():
    rows = query("""SELECT i.name,i.manufacturer,i.model,i.serial,i.sku,c.name,pt.name,
                           i.condition,i.owner_company,i.purchase_date,i.shelf,i.qty,
                           COALESCE(i.qty_out,0),i.checked_out,i.checkout_by,i.job_ref,
                           i.cost_price,i.sale_price,i.tax_paid,i.tax_rate,i.low_stock_threshold,
                           i.cpu,i.ram,i.storage,i.notes,i.created_at
                    FROM items i
                    LEFT JOIN categories c ON c.id=i.category_id
                    LEFT JOIN product_types pt ON pt.id=i.product_type_id
                    WHERE i.active=1 ORDER BY i.name""")
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Name","Manufacturer","Model","Serial","SKU","Category","Product Type",
                "Condition","Owner/Company","Purchase Date","Shelf","Qty","Qty Out",
                "Checked Out","Checked Out By","Job Ref","Cost ($)","Sale ($)",
                "Tax Paid","Tax Rate (%)","Low Stock Threshold","CPU","RAM","Storage","Notes","Created At"])
    for r in rows:
        row=list(r); row[13]="Yes" if row[13] else "No"; row[18]={1:"Yes",0:"No"}.get(row[18],"Not Set")
        w.writerow(row)
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype="text/csv",
                     as_attachment=True, download_name=f"storelax_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")

@app.route("/export/excel")
@login_required
@perm_required("import_export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500
    rows = query("""SELECT i.name,i.manufacturer,i.model,i.serial,i.sku,c.name,pt.name,
                           i.condition,i.owner_company,i.purchase_date,i.shelf,i.qty,
                           COALESCE(i.qty_out,0),i.checked_out,i.checkout_by,i.job_ref,
                           i.cost_price,i.sale_price,i.tax_paid,i.tax_rate,i.low_stock_threshold,
                           i.cpu,i.ram,i.storage,i.notes,i.created_at
                    FROM items i
                    LEFT JOIN categories c ON c.id=i.category_id
                    LEFT JOIN product_types pt ON pt.id=i.product_type_id
                    WHERE i.active=1 ORDER BY c.name,i.name""")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Inventory"
    hdrs=["Name","Manufacturer","Model","Serial","SKU","Category","Product Type",
          "Condition","Owner/Company","Purchase Date","Shelf","Qty Total","Qty Out",
          "Checked Out","Checked Out By","Job Ref","Cost ($)","Sale ($)",
          "Tax Paid","Tax Rate (%)","Low Stock Threshold","CPU","RAM","Storage","Notes","Created At"]
    hfill=PatternFill("solid",fgColor="0F172A"); hfont=Font(color="FFFFFF",bold=True,size=10)
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=1,column=col,value=h); c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal="center")
    for i,w in enumerate([28,16,16,20,12,14,16,12,18,12,8,8,8,12,14,18,10,10,10,10,10,12,12,12,28,18],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    for ri,r in enumerate(rows,2):
        vals=list(r); vals[13]="Yes" if vals[13] else "No"; vals[18]={1:"Yes",0:"No"}.get(vals[18],"Not Set")
        for ci,v in enumerate(vals,1): ws.cell(row=ri,column=ci,value=v)
    ws2=wb.create_sheet("Summary")
    total=query("SELECT COUNT(*) FROM items WHERE active=1",one=True)[0]
    out_c=query("SELECT COUNT(*) FROM items WHERE active=1 AND checked_out=1",one=True)[0]
    low_c=query("SELECT COUNT(*) FROM items WHERE active=1 AND qty IS NOT NULL AND low_stock_threshold>0 AND (qty-COALESCE(qty_out,0))<=low_stock_threshold",one=True)[0]
    tcost=query("SELECT SUM(cost_price) FROM items WHERE active=1",one=True)[0] or 0
    tsale=query("SELECT SUM(sale_price) FROM items WHERE active=1 AND sale_price IS NOT NULL",one=True)[0] or 0
    ws2.append(["Metric","Value"])
    for row in [("Total Items",total),("Checked Out",out_c),("Available",total-out_c),
                ("Low Stock Alerts",low_c),("Total Cost Value",round(tcost,2)),
                ("Total Sale Value",round(tsale,2)),("Potential Profit",round(tsale-tcost,2)),
                ("Export Date",datetime.now().strftime("%Y-%m-%d %H:%M"))]:
        ws2.append(list(row))
    output=io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"storelax_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@app.route("/import/excel", methods=["POST"])
@login_required
@admin_required
def import_excel():
    try: import openpyxl
    except ImportError: return jsonify({"ok":False,"msg":"openpyxl not installed"})
    f = request.files.get("file")
    if not f: return jsonify({"ok":False,"msg":"No file uploaded"})
    try:
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
        if len(rows)<2: return jsonify({"ok":False,"msg":"File is empty"})
        cat_map={r["name"].lower():r["id"] for r in query("SELECT id,name FROM categories")}
        pt_map={r["name"].lower():r["id"] for r in query("SELECT id,name FROM product_types WHERE active=1")}
        now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); added=0; skipped=0; errors=[]
        for i,row in enumerate(rows[1:],2):
            try:
                name=str(row[0] or "").strip()
                if not name: skipped+=1; continue
                d=dict(name=name,manufacturer=str(row[1] or "").strip() or None,
                       model=str(row[2] or "").strip() or None,serial=str(row[3] or "").strip() or None,
                       sku=str(row[4] or "").strip() or None,
                       category_id=cat_map.get(str(row[5] or "").strip().lower()),
                       product_type_id=pt_map.get(str(row[6] or "").strip().lower()),
                       condition=str(row[7] or "New").strip(),
                       owner_company=str(row[8] or "").strip() or None,
                       purchase_date=str(row[9] or "").strip() or None,
                       shelf=str(row[10] or "").strip() or None,
                       qty=row[11],tax_paid=-1,tax_rate=row[19] or 0,
                       low_stock_threshold=int(row[20]) if row[20] else 0,
                       cpu=str(row[21] or "").strip() or None,ram=str(row[22] or "").strip() or None,
                       storage=str(row[23] or "").strip() or None,notes=str(row[24] or "").strip(),
                       cost_price=row[16],sale_price=row[17])
                _save_item(d); log_action("ITEM_IMPORT",detail=f"Imported: {name}"); added+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        return jsonify({"ok":True,"added":added,"skipped":skipped,"errors":errors[:10]})
    except Exception as e: return jsonify({"ok":False,"msg":str(e)})

if __name__ == "__main__":
    seed_db()
    print("\n Storelax running at http://localhost:5000")
    print("   Default logins: admin/admin123  and  worker/worker123\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
